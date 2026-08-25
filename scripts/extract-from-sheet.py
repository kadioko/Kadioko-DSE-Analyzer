"""
Extracts import-ready CSVs from the Kadioko DSE Sheet workbook.

    pip install openpyxl
    python scripts/extract-from-sheet.py "docs/Kadioko DSE Sheet.xlsx" out/

Produces three files that the Analyzer's own importers consume:

    instruments.csv   security master (symbol, name, sector, shares outstanding)
    market.csv        daily observations for /admin/data (kind=market)
    fundamentals.csv  reported financial figures for /admin/data (kind=fundamentals)

This is a ONE-WAY migration helper, not a permanent dependency. The Sheet is
intended to become a consumer of the Analyzer's API rather than its source; this
script exists to carry existing data across once.

Nothing is invented. Every value written comes from a cell in the workbook. The
Sheet's own Fundamental Score is deliberately NOT copied: the Analyzer computes
its own score from the raw reported figures, so the whole pipeline is
reproducible from source data rather than trusting a number typed into a
spreadsheet. The two scores will differ, and that is expected.
"""

from __future__ import annotations

import csv
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")


def as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def iso(value) -> str:
    d = as_date(value)
    return d.isoformat() if d else ""


def num(value) -> str:
    """Numbers pass through; anything else becomes an EMPTY cell, never a 0."""
    if value is None or isinstance(value, str):
        return ""
    if isinstance(value, (int, float)):
        return repr(value) if isinstance(value, float) else str(value)
    return ""


def header_index(rows: list[tuple], needle: str) -> int | None:
    """Finds the row index whose first cell matches a known header label."""
    for i, row in enumerate(rows):
        if row and isinstance(row[0], str) and row[0].strip() == needle:
            return i
    return None


def period_type_for(statement: date, cadence: str) -> str:
    """Maps the Sheet's cadence plus a statement date onto a period type."""
    cadence = (cadence or "").strip().lower()
    if cadence == "full":
        return "FY"
    if cadence == "half":
        return "H1" if statement.month <= 6 else "H2"
    if cadence == "quarter":
        return {3: "Q1", 6: "Q2", 9: "Q3", 12: "Q4"}.get(statement.month, "INTERIM")
    return "FY"


def prior_period_end(statement: date) -> date:
    """
    The comparable period one year earlier.

    The Sheet carries prior-year comparatives inline. Writing them as their own
    row gives the Analyzer's engine a real baseline to compute growth against,
    rather than having growth silently unavailable.
    """
    try:
        return statement.replace(year=statement.year - 1)
    except ValueError:  # 29 February
        return statement.replace(year=statement.year - 1, day=28)


def extract(workbook_path: str, out_dir: str) -> None:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    # ---------------------------------------------------------------- symbols
    rows = list(wb["Symbols"].iter_rows(values_only=True))
    start = header_index(rows, "Symbol")
    if start is None:
        sys.exit('Could not find the "Symbol" header row in the Symbols sheet.')
    head = [str(c or "").strip() for c in rows[start]]
    col = {name: i for i, name in enumerate(head)}

    instruments = []
    for row in rows[start + 1 :]:
        symbol = (row[col["Symbol"]] or "") if col.get("Symbol") is not None else ""
        if not isinstance(symbol, str) or not symbol.strip():
            continue
        instruments.append(
            {
                "symbol": symbol.strip().upper(),
                "name": str(row[col.get("Company / Security Name", 1)] or "").strip(),
                "sector": str(row[col.get("Sector", 3)] or "").strip(),
                "currency": str(row[col.get("Currency", 6)] or "TZS").strip(),
                "shares_outstanding": num(row[col.get("Shares Outstanding", 7)]),
                "active": "true"
                if str(row[col.get("Active", 5)] or "Yes").strip().lower()
                in {"yes", "true", "y"}
                else "false",
            }
        )

    with (out / "instruments.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(instruments[0].keys()))
        w.writeheader()
        w.writerows(instruments)

    # ----------------------------------------------------------------- market
    rows = list(wb["Daily Data"].iter_rows(values_only=True))
    start = header_index(rows, "Date")
    if start is None:
        sys.exit('Could not find the "Date" header row in the Daily Data sheet.')
    head = [str(c or "").strip() for c in rows[start]]
    col = {name: i for i, name in enumerate(head)}

    def cell(row, name):
        i = col.get(name)
        return row[i] if i is not None and i < len(row) else None

    market = []
    for row in rows[start + 1 :]:
        d = as_date(cell(row, "Date"))
        symbol = cell(row, "Symbol")
        if not d or not isinstance(symbol, str) or not symbol.strip():
            continue
        market.append(
            {
                "Date": d.isoformat(),
                "Symbol": symbol.strip().upper(),
                # The workbook's "Open" column holds the PREVIOUS session's
                # close, not an opening price: across the whole Daily Data
                # sheet it matches the prior close on 118 of 118 comparable
                # rows, and it routinely sits outside the same day's high/low.
                # It is therefore mapped to Previous Close, which the sheet
                # leaves empty. Open is exported blank rather than guessed.
                "Open": "",
                "Previous Close": num(cell(row, "Previous Close"))
                or num(cell(row, "Open")),
                "Close": num(cell(row, "Close")),
                "High": num(cell(row, "High")),
                "Low": num(cell(row, "Low")),
                "Change": num(cell(row, "Change %")),
                "Turnover": num(cell(row, "Turnover TZS")),
                "Deals": num(cell(row, "Deals")),
                "Volume": num(cell(row, "Volume")),
                "Outstanding Bid": num(cell(row, "Outstanding Bid Qty")),
                "Outstanding Offer": num(cell(row, "Outstanding Offer Qty")),
                "Market Cap": num(cell(row, "Market Cap TZS")),
            }
        )

    market.sort(key=lambda r: (r["Date"], r["Symbol"]))
    with (out / "market.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(market[0].keys()))
        w.writeheader()
        w.writerows(market)

    # ----------------------------------------------------------- fundamentals
    rows = list(wb["Fundamental Scorecard"].iter_rows(values_only=True))
    start = header_index(rows, "Ticker")
    if start is None:
        sys.exit('Could not find the "Ticker" header row in the Fundamental Scorecard.')
    head = [str(c or "").strip() for c in rows[start]]
    col = {name: i for i, name in enumerate(head)}

    def fcell(row, *names):
        for name in names:
            i = col.get(name)
            if i is not None and i < len(row) and row[i] is not None:
                return row[i]
        return None

    fundamentals = []
    for row in rows[start + 1 :]:
        ticker = fcell(row, "Ticker")
        if not isinstance(ticker, str) or not ticker.strip():
            continue
        statement = as_date(fcell(row, "Statement Date"))
        if not statement:
            continue

        cadence = str(fcell(row, "Period Type") or "full")
        ptype = period_type_for(statement, cadence)
        # NOT "Last Updated". That column is the sheet's own maintenance stamp -
        # identical on every row, and dated in the future - not the date the
        # issuer released the results. Mapping it to published_at made the
        # no-look-ahead rule discard perfectly good figures for CRDB, NMB, TBL
        # and TPCC, because results cannot be used before they were published.
        # No column in this workbook carries a real publication date, so none is
        # claimed; the as-of rule then falls back to the period end.
        published = None

        revenue = fcell(
            row,
            "Revenue — current period",
            "Revenue – current period",
            "Revenue - current period",
        )
        revenue_prior = fcell(
            row,
            "Revenue — prior year",
            "Revenue – prior year",
            "Revenue - prior year",
        )
        profit = fcell(
            row,
            "Profit after tax — current",
            "Profit after tax — current period",
            "Profit after tax - current period",
        )
        profit_prior = fcell(
            row,
            "Profit after tax — prior year",
            "Profit after tax – prior year",
            "Profit after tax - prior year",
        )
        equity = fcell(row, "Total equity")
        debt = fcell(row, "Borrowings")

        fundamentals.append(
            {
                "symbol": ticker.strip().upper(),
                "period_end": statement.isoformat(),
                "period_type": ptype,
                "published_at": published.isoformat() if published else "",
                "revenue": num(revenue),
                "net_income": num(profit),
                "total_equity": num(equity),
                "total_debt": num(debt),
                "source": "Kadioko DSE Sheet - Fundamental Scorecard",
                "verified": "",
            }
        )

        # The prior-year comparative from the same filing, as its own row, so
        # growth has a real baseline instead of being unavailable.
        if revenue_prior is not None or profit_prior is not None:
            fundamentals.append(
                {
                    "symbol": ticker.strip().upper(),
                    "period_end": prior_period_end(statement).isoformat(),
                    "period_type": ptype,
                    "published_at": published.isoformat() if published else "",
                    "revenue": num(revenue_prior),
                    "net_income": num(profit_prior),
                    "total_equity": "",
                    "total_debt": "",
                    "source": "Kadioko DSE Sheet - prior-year comparative",
                    "verified": "",
                }
            )

    # Oldest first, so the prior period is stored before the one that needs it.
    fundamentals.sort(key=lambda r: (r["symbol"], r["period_end"]))
    with (out / "fundamentals.csv").open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=list(fundamentals[0].keys()))
        w.writeheader()
        w.writerows(fundamentals)

    print(f"instruments.csv   {len(instruments)} rows")
    print(f"market.csv        {len(market)} rows")
    print(f"fundamentals.csv  {len(fundamentals)} rows")
    print(f"\nWritten to {out.resolve()}")
    print("\nImport order: instruments (db:seed) -> market -> fundamentals")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    extract(sys.argv[1], sys.argv[2])
